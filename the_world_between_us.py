# coding=utf-8
import requests
import time
import re
import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import jieba
from wordcloud import WordCloud


# 复制浏览器中的request headers
headers = {
    'Accept': 'application/json, text/plain, */*',
    # Accept-Encoding不能留，否则返回的文件无法解码
    # 'Accept-Encoding': 'gzip, deflate, br',
}
# 复制浏览器中的cookies
cookies = {

}
drama_code = 30181230
base_url = "https://movie.douban.com/subject/{}/comments".format(drama_code)


def get_comments():
    title = ['用户名', '是否看过', '星级', '评论内容', '赞同数', '评论日期', '评论时间']
    file_name = 'the_word_comments.xlsx'
    write_excel(title, file_name)
    for i in range(25):
        time.sleep(1)
        params = {'percent_type': '', 'start': str(20*i), 'limit': '20', 'status': 'P', 'sort': 'new_score',
                  'comments_only': '1', 'ck': 'qN8_'}
        try:
            # 发送GET请求获取数据,headers和cookies从浏览器中获取
            response = requests.get(base_url, headers=headers, cookies=cookies, params=params)
            if response.status_code != 200:
                break
            result = response.json()
            # 用正则匹配出需要的数据
            user_name = re.findall('<a title="(.*?)"', result['html'])
            seen = re.findall('<span>(.*?)</span>', result['html'])
            star = re.findall('rating" title="(.*?)"></span>', result['html'])
            comment = re.findall('<span class="short">(.*?)</span>', result['html'], re.S)
            like = re.findall(r'<span class="votes vote-count">(\d+?)</span>', result['html'])
            comment_time = re.findall('<span class="comment-time " title="(.*?)">', result['html'])
            # 部分数可能不完整，用控制补充缺失的数据，使每个数据列表的长度相等
            data_full([user_name, seen, star, comment, like, comment_time])
            for j in range(20):
                # 将数据保存到excel中
                data = [user_name[j], seen[j], star[j], comment[j], like[j], *comment_time[j].split(' ')]
                write_excel(data, file_name)
            print('[INFO]第{}页数据获取成功。'.format(i + 1, ))
        except Exception as e:
            print('[ERROR]第{}页数据获取失败:{}'.format(i + 1, e))


def data_full(label):
    for key in label:
        if len(key) < 20:
            for _ in range(20 - len(key)):
                key.append('')


def write_excel(data, file_name):
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        rows = workbook.active.max_row
    else:
        workbook = openpyxl.Workbook()
        rows = 0
    ws = workbook.active
    for i in range(len(data)):
        ws.cell(row=rows + 1, column=i + 1).value = data[i]
    workbook.save(filename=file_name)


def is_user_repeat():
    """判断是否有重复用户"""
    df = pd.read_excel('the_word_comments.xlsx')
    user = df['用户名']
    print(user[user.duplicated().isin(['True'])])


def is_all_seen():
    """判断是否所有用户都看过"""
    df = pd.read_excel('the_word_comments.xlsx')
    seen = df['是否看过']
    seen.value_counts().plot.pie(figsize=(5, 2.4), ylabel='', autopct='%.0f%%', colors=['c'],
                                 textprops={'fontsize': 24, 'color': 'b'}, pctdistance=0, labeldistance=1.5)
    plt.show()


def star_distribution():
    """星级分布"""
    df = pd.read_excel('the_word_comments.xlsx')
    star = df['星级']
    five, four, three, one, two, zero = star.value_counts(dropna=False)
    star_num = {'0星': zero, '1星': one, '2星': two, '3星': three, '4星': four, '5星': five}
    fig = pd.Series(star_num).plot.barh(figsize=(5, 3), color='#FF9900', width=0.7)
    for a, b in zip(range(len(star_num)), star_num.values()):
        plt.text(b+2, a, '%.0f' % b, ha='left', va='center')
    plt.xticks(range(0, 400, 50))
    # 隐藏x标签
    x_ticks = fig.xaxis.get_major_ticks()
    for i in range(len(x_ticks)):
        x_ticks[i].set_visible(False)
    # 隐藏边框线
    for spine in fig.spines:
        fig.spines[spine].set_visible(False)
    # 隐藏刻度线
    fig.tick_params(bottom=False, top=False, left=False, right=False)
    plt.show()


def gen_word_cloud():
    """生成词云图"""
    df = pd.read_excel('the_word_comments.xlsx')
    comment = df['评论内容']
    for i in range(500):
        with open('comment.txt', 'a', encoding='utf-8') as f:
            f.write(comment[i])
    with open('comment.txt', 'r', encoding='utf-8') as f:
        all_comment = f.read()
    cut_text = jieba.cut(all_comment)
    result = ' '.join(cut_text)
    exclude = {"我", "我们", "与", "恶", "恶的", "的", "距离", "每", "一个", "不", "知道", "这", "这个", "而", "是", "不是",
               "都是", "也是", "是我", "让我", "都在", "都是", "也是", "都", "也", "是", "看完", "还在", "这部", "剧", "还",
               "在", "看", "完", "更多", "更", "多", "最后", "一集", "让人", "让", "人", "并", "没有", "了", "很", "好"}
    wc = WordCloud(font_path="simhei.ttf", width=400, height=300, max_words=500, min_font_size=4,
                   background_color='white', contour_color='black', stopwords=exclude)
    wc.generate(result)
    wc.to_file("comment.png")


def hot_word_count():
    """高频词出现的次数"""
    with open('comment.txt', 'r', encoding='utf-8') as f:
        all_comment = f.read()
    cut_text = jieba.cut(all_comment)
    result = ' '.join(cut_text)
    word_count = {
        '民主': len(re.findall('民主', result)),
        '法治': len(re.findall('法治', result)),
        '社会': len(re.findall('社会', result)),
        '受害者': len(re.findall('受害者', result)),
        '家属': len(re.findall('家属', result)),
        '精神': len(re.findall('精神', result)),
        '新闻': len(re.findall('新闻', result)),
        '律师': len(re.findall('律师', result)),
        '最佳': len(re.findall('新闻', result))
    }
    word_count_df = {
        'index': list(reversed([name for name in word_count.keys()])),
        'count': list(reversed([count for count in word_count.values()])),
        'x': ['count' for _ in range(len(word_count))],
        'y': [(i + 1) * 100 for i in range(len(word_count))]
    }
    df = pd.DataFrame(word_count_df)
    fig = df.plot.scatter('x', 'y', figsize=(4, 10), s='count', c='#FF9900', ylabel='', xlabel='')
    plt.yticks(word_count_df['y'], word_count_df['index'])
    plt.xticks([])
    for a, b, c in zip([0]*len(word_count_df['x']), word_count_df['y'], word_count_df['count']):
        plt.text(a+0.05, b, '%.0f' % c, ha='center', va='center', fontdict={'size': 16})
    for spine in fig.spines:
        fig.spines[spine].set_visible(False)
    fig.tick_params(axis='y', colors='b', labelsize=16)
    fig.tick_params(bottom=False, top=False, left=False, right=False)
    plt.show()


def likes_distribution():
    """评论点赞数分布"""
    df = pd.read_excel('the_word_comments.xlsx')
    likes = df['赞同数']
    likes_division = {
        '少于100': len(likes[(0 <= likes) & (likes < 100)]),
        '100至500': len(likes[(100 <= likes) & (likes < 500)]),
        '超过500': len(likes[500 <= likes])
    }
    explode = (0, 0.05, 0.1)
    pd.Series(likes_division).plot.pie(figsize=(5, 3), ylabel='', autopct='%.0f%%', pctdistance=0.85,
                                       colors=['c', 'b', '#FF9900'],
                                       explode=explode, startangle=7.5, textprops={'fontsize': 14, 'color': 'm'})
    plt.pie([i for i in likes_division.values()], radius=0.7, colors='w', explode=explode, startangle=7.5)
    plt.pie([1], radius=0.7, colors='w')
    plt.show()


def hot_comment_likes():
    """热评的点赞数"""
    df = pd.read_excel('the_word_comments.xlsx')
    top_10_likes = df['赞同数'].sort_values(ascending=False).head(10)
    fig = top_10_likes.plot.bar(figsize=(5, 3), color='#0099CC', width=0.7, )
    for a, b in zip(range(len(top_10_likes)), top_10_likes):
        plt.text(a, b+100, '%.0f' % b, ha='center', va='bottom')
    plt.xticks(range(10), ['Hot{}'.format(i+1) for i in range(10)], rotation=0)
    plt.yticks([])
    fig.tick_params(axis='x', colors='#FF0033')
    for spine in fig.spines:
        fig.spines[spine].set_visible(False)
    fig.tick_params(bottom=False, top=False, left=False, right=False)
    plt.show()


if __name__ == '__main__':
    # get_comments()
    # is_user_repeat()
    # is_all_seen()
    # 数据分析
    # star_distribution()
    # gen_word_cloud()
    # hot_word_count()
    # likes_distribution()
    hot_comment_likes()
